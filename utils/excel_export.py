"""
Excel Export utility for PlywoodPro.
Uses openpyxl for GSTR-1 and ITC Register exports.
"""

import os
import traceback
from pathlib import Path
import tkinter as tk
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


EXPORT_DIR = Path(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))) / "exports"
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
HEADER_FILL = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
TOTAL_FONT = Font(bold=True, size=10)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'))


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)


def _add_headers(ws, headers):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER


def _fmt(val):
    try: return round(float(val), 2)
    except: return val or ''


def _ask_save_location(default_filename: str, filetype: str = "xlsx") -> str | None:
    """Show file save dialog. Returns chosen path or None if cancelled."""
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)

    if filetype == "pdf":
        filetypes = [("PDF Files", "*.pdf"), ("All Files", "*.*")]
        ext = ".pdf"
    else:
        filetypes = [("Excel Files", "*.xlsx"), ("All Files", "*.*")]
        ext = ".xlsx"

    initial_dir = os.path.expanduser("~/Documents")
    path = filedialog.asksaveasfilename(
        title="Save File",
        initialdir=initial_dir,
        initialfile=default_filename,
        defaultextension=ext,
        filetypes=filetypes,
        parent=root
    )
    root.destroy()
    return path if path else None


def export_gstr1_excel(month, year):
    """Export GSTR-1 with 4 sheets: B2B, B2C, CDNR, HSN Summary."""
    from modules.gst import get_gstr1_b2b, get_gstr1_b2c, get_gstr1_cdnr, get_gstr1_hsn_summary

    try:
        EXPORT_DIR.mkdir(exist_ok=True)
        wb = Workbook()

        # Sheet 1: B2B
        ws = wb.active
        ws.title = "B2B"
        _add_headers(ws, ['GSTIN', 'Party', 'Invoice No', 'Date', 'Taxable', 'CGST', 'SGST', 'IGST', 'Total', 'Rate%'])
        b2b = get_gstr1_b2b(month, year)
        for r in b2b:
            ws.append([r.get('gstin',''), r.get('party_name',''), r.get('voucher_no',''),
                       r.get('date',''), _fmt(r.get('taxable_amount')),
                       _fmt(r.get('cgst_amount')), _fmt(r.get('sgst_amount')),
                       _fmt(r.get('igst_amount')), _fmt(r.get('total_amount')),
                       _fmt(r.get('gst_rate'))])
        if b2b:
            ws.append(['', '', '', 'TOTAL',
                sum(_fmt(r.get('taxable_amount',0)) for r in b2b),
                sum(_fmt(r.get('cgst_amount',0)) for r in b2b),
                sum(_fmt(r.get('sgst_amount',0)) for r in b2b),
                sum(_fmt(r.get('igst_amount',0)) for r in b2b),
                sum(_fmt(r.get('total_amount',0)) for r in b2b), ''])
            for cell in ws[ws.max_row]: cell.font = TOTAL_FONT
        _auto_width(ws)

        # Sheet 2: B2C
        ws2 = wb.create_sheet("B2C")
        _add_headers(ws2, ['State', 'GST Rate%', 'Taxable', 'CGST', 'SGST', 'IGST', 'Total'])
        b2c = get_gstr1_b2c(month, year)
        for r in b2c:
            ws2.append([r.get('state',''), _fmt(r.get('gst_rate')),
                        _fmt(r.get('taxable_value')), _fmt(r.get('cgst')),
                        _fmt(r.get('sgst')), _fmt(r.get('igst')), _fmt(r.get('total'))])
        _auto_width(ws2)

        # Sheet 3: CDNR
        ws3 = wb.create_sheet("CDNR")
        _add_headers(ws3, ['GSTIN', 'Party', 'Note No', 'Date', 'Ref', 'Taxable', 'CGST', 'SGST', 'IGST', 'Total'])
        cdnr = get_gstr1_cdnr(month, year)
        for r in cdnr:
            ws3.append([r.get('gstin',''), r.get('party_name',''), r.get('voucher_no',''),
                        r.get('date',''), r.get('reference_no',''),
                        _fmt(r.get('taxable_amount')), _fmt(r.get('cgst_amount')),
                        _fmt(r.get('sgst_amount')), _fmt(r.get('igst_amount')),
                        _fmt(r.get('total_amount'))])
        _auto_width(ws3)

        # Sheet 4: HSN Summary
        ws4 = wb.create_sheet("HSN Summary")
        _add_headers(ws4, ['HSN Code', 'Description', 'Unit', 'Qty', 'Value', 'Taxable', 'CGST', 'SGST', 'IGST'])
        hsn = get_gstr1_hsn_summary(month, year)
        for r in hsn:
            ws4.append([r.get('hsn_code',''), r.get('description',''), r.get('unit',''),
                        _fmt(r.get('total_qty')), _fmt(r.get('total_value')),
                        _fmt(r.get('total_taxable')), _fmt(r.get('cgst')),
                        _fmt(r.get('sgst')), _fmt(r.get('igst'))])
        if hsn:
            ws4.append(['', '', 'TOTAL', sum(_fmt(r.get('total_qty',0)) for r in hsn),
                sum(_fmt(r.get('total_value',0)) for r in hsn),
                sum(_fmt(r.get('total_taxable',0)) for r in hsn),
                sum(_fmt(r.get('cgst',0)) for r in hsn),
                sum(_fmt(r.get('sgst',0)) for r in hsn),
                sum(_fmt(r.get('igst',0)) for r in hsn)])
            for cell in ws4[ws4.max_row]: cell.font = TOTAL_FONT
        _auto_width(ws4)

        filename = f"GSTR1_{year}_{month:02d}.xlsx"
        path = _ask_save_location(filename, "xlsx")
        if not path:
            return False, "Export cancelled."
        wb.save(path)
        return True, path

    except Exception as e:
        traceback.print_exc()
        return False, str(e)


def export_itc_register_excel(date_from='', date_to=''):
    """Export ITC Register to Excel."""
    from modules.gst import get_itc_register

    try:
        EXPORT_DIR.mkdir(exist_ok=True)
        data = get_itc_register(date_from, date_to)
        wb = Workbook()
        ws = wb.active
        ws.title = "ITC Register"

        _add_headers(ws, ['Invoice', 'Date', 'Party', 'GSTIN', 'Description', 'HSN',
                          'Taxable', 'CGST', 'SGST', 'IGST', 'Total', 'Status'])
        for r in data['entries']:
            ws.append([r.get('voucher_no',''), r.get('date',''), r.get('party_name',''),
                       r.get('gstin',''), r.get('description',''), r.get('hsn_code',''),
                       _fmt(r.get('taxable_amount')), _fmt(r.get('cgst_amount')),
                       _fmt(r.get('sgst_amount')), _fmt(r.get('igst_amount')),
                       _fmt(r.get('total_amount')), r.get('status','')])

        ws.append(['', '', '', '', '', 'ELIGIBLE TOTAL', '',
                   data['total_eligible_cgst'], data['total_eligible_sgst'],
                   data['total_eligible_igst'], data['total_eligible_itc'], ''])
        for cell in ws[ws.max_row]: cell.font = TOTAL_FONT
        _auto_width(ws)

        path = _ask_save_location("ITC_Register.xlsx", "xlsx")
        if not path:
            return False, "Export cancelled."
        wb.save(path)
        return True, path

    except Exception as e:
        traceback.print_exc()
        return False, str(e)


def export_sales_register_excel(date_from='', date_to=''):
    """Export Sales Register to Excel."""
    from modules.reports import get_sales_register
    try:
        EXPORT_DIR.mkdir(exist_ok=True)
        rows = get_sales_register(date_from, date_to)
        wb = Workbook(); ws = wb.active; ws.title = "Sales Register"
        _add_headers(ws, ['Date', 'Invoice', 'Party', 'Taxable', 'CGST', 'SGST', 'IGST', 'Total'])
        for r in rows:
            ws.append([r.get('date',''), r.get('voucher_no',''), r.get('party_name',''),
                       _fmt(r.get('taxable_amount')), _fmt(r.get('cgst_amount')),
                       _fmt(r.get('sgst_amount')), _fmt(r.get('igst_amount')),
                       _fmt(r.get('grand_total'))])
        if rows:
            ws.append(['','','TOTAL',
                sum(_fmt(r.get('taxable_amount',0)) for r in rows),
                sum(_fmt(r.get('cgst_amount',0)) for r in rows),
                sum(_fmt(r.get('sgst_amount',0)) for r in rows),
                sum(_fmt(r.get('igst_amount',0)) for r in rows),
                sum(_fmt(r.get('grand_total',0)) for r in rows)])
            for cell in ws[ws.max_row]: cell.font = TOTAL_FONT
        _auto_width(ws)
        path = _ask_save_location("Sales_Register.xlsx", "xlsx")
        if not path:
            return False, "Export cancelled."
        wb.save(path)
        return True, path
    except Exception as e:
        traceback.print_exc()
        return False, str(e)


def export_party_outstanding_excel(party_type=''):
    """Export Party Outstanding to Excel."""
    from modules.reports import get_party_outstanding
    try:
        EXPORT_DIR.mkdir(exist_ok=True)
        rows = get_party_outstanding(party_type)
        wb = Workbook(); ws = wb.active; ws.title = "Outstanding"
        _add_headers(ws, ['Party', 'Type', 'Invoiced', 'Paid', 'Balance', 'Oldest Invoice'])
        for r in rows:
            ws.append([r.get('party_name',''), r.get('type',''),
                       _fmt(r.get('total_invoiced')), _fmt(r.get('total_paid')),
                       _fmt(r.get('balance')), r.get('oldest_invoice_date','')])
        _auto_width(ws)
        path = _ask_save_location("Party_Outstanding.xlsx", "xlsx")
        if not path:
            return False, "Export cancelled."
        wb.save(path)
        return True, path
    except Exception as e:
        traceback.print_exc()
        return False, str(e)


def export_item_profit_excel():
    """Export Item Profit Report to Excel."""
    from modules.reports import get_item_profit_report
    try:
        EXPORT_DIR.mkdir(exist_ok=True)
        rows = get_item_profit_report()
        wb = Workbook(); ws = wb.active; ws.title = "Item Profit"
        _add_headers(ws, ['Item', 'HSN', 'Purch Qty', 'Avg Purch Rate', 'Sold Qty',
                          'Avg Sale Rate', 'Gross Profit', 'Margin%'])
        for r in rows:
            ws.append([r.get('item_name',''), r.get('hsn_code',''),
                       _fmt(r.get('purchased_qty')), _fmt(r.get('avg_purchase_rate')),
                       _fmt(r.get('sold_qty')), _fmt(r.get('avg_sale_rate')),
                       _fmt(r.get('gross_profit')), _fmt(r.get('margin_pct'))])
        _auto_width(ws)
        path = _ask_save_location("Item_Profit.xlsx", "xlsx")
        if not path:
            return False, "Export cancelled."
        wb.save(path)
        return True, path
    except Exception as e:
        traceback.print_exc()
        return False, str(e)
