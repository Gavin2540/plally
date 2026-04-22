"""
Phase 6 Integration Test -- PlywoodPro Dashboard & Reports
Runs after Phase 2 seed data.
"""
import os, sys
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

from modules.reports import (
    get_dashboard_stats, get_sales_register, get_purchase_register,
    get_party_outstanding, get_item_profit_report, get_stock_valuation_report,
)
from utils.pdf_export import export_party_outstanding_pdf
from utils.excel_export import export_sales_register_excel

PASS = FAIL = 0

def check(label, cond, detail=""):
    global PASS, FAIL
    if cond:
        PASS += 1; print(f"  [PASS] {label}")
    else:
        FAIL += 1; print(f"  [FAIL] {label} -- {detail}")

def main():
    global PASS, FAIL
    print("=" * 60)
    print("PlywoodPro Phase 6 -- Dashboard & Reports Integration Test")
    print("=" * 60)

    # ── 1. Dashboard Stats ──
    print("\n[1] Dashboard Stats")
    stats = get_dashboard_stats()
    check("Dashboard returns data", isinstance(stats, dict))
    check("today_sales key exists", 'today_sales' in stats)
    check("receivables key exists", 'receivables' in stats)
    check("low_stock_count key exists", 'low_stock_count' in stats)
    check("recent_invoices is list", isinstance(stats.get('recent_invoices'), list))
    check("Recent invoices found", len(stats['recent_invoices']) > 0,
          f"Got {len(stats['recent_invoices'])}")
    print(f"    Today's Sales: {stats['today_sales']}")
    print(f"    Receivables:   {stats['receivables']}")
    print(f"    Low Stock:     {stats['low_stock_count']}")

    # ── 2. Sales Register ──
    print("\n[2] Sales Register")
    sr = get_sales_register()
    check("Sales register has entries", len(sr) > 0, f"Got {len(sr)}")
    if sr:
        r = sr[0]
        check("Has invoice_no", bool(r.get('voucher_no')))
        check("Has party_name", bool(r.get('party_name')))
        check("Grand total > 0", (r.get('grand_total') or 0) > 0)
        print(f"    First: {r['voucher_no']} | {r['party_name']} | {r['grand_total']}")

    # ── 3. Purchase Register ──
    print("\n[3] Purchase Register")
    pr = get_purchase_register()
    check("Purchase register has entries", len(pr) > 0, f"Got {len(pr)}")

    # ── 4. Party Outstanding ──
    print("\n[4] Party Outstanding")
    po = get_party_outstanding()
    check("Party outstanding has entries", len(po) > 0, f"Got {len(po)}")
    chennai = next((p for p in po if 'Chennai' in p.get('party_name', '')), None)
    check("Chennai Woodworks found", chennai is not None)
    if chennai:
        check("Chennai balance > 0", chennai['balance'] > 0, f"Balance: {chennai['balance']}")
        print(f"    Chennai: invoiced={chennai['total_invoiced']}, paid={chennai['total_paid']}, balance={chennai['balance']}")

    # ── 5. Item Profit Report ──
    print("\n[5] Item Profit Report")
    ipr = get_item_profit_report()
    check("Item profit has entries", len(ipr) > 0, f"Got {len(ipr)}")
    # Find the item that was actually sold in Phase 2
    sold_item = next((r for r in ipr if (r.get('sold_qty') or 0) > 0), None)
    check("Found a sold item", sold_item is not None)
    if sold_item:
        check("Has item_name", bool(sold_item.get('item_name')))
        check("Has margin_pct", 'margin_pct' in sold_item)
        check("Avg purchase rate > 0", (sold_item.get('avg_purchase_rate') or 0) > 0)
        check("Avg sale rate > 0", (sold_item.get('avg_sale_rate') or 0) > 0)
        print(f"    {sold_item['item_name']}: buy={sold_item['avg_purchase_rate']}, sell={sold_item['avg_sale_rate']}, margin={sold_item['margin_pct']}%")

    # ── 6. Stock Valuation ──
    print("\n[6] Stock Valuation")
    sv = get_stock_valuation_report()
    check("Stock valuation has entries", len(sv) > 0, f"Got {len(sv)}")
    if sv:
        total_val = sum(r['stock_value'] for r in sv)
        check("Total stock value > 0", total_val > 0, f"Total: {total_val}")
        print(f"    Total stock value: {total_val}")

    # ── 7. Export Party Outstanding PDF ──
    print("\n[7] Export Party Outstanding PDF")
    ok, path = export_party_outstanding_pdf()
    check("PDF exported", ok, path if not ok else "")
    if ok:
        check("PDF file exists", os.path.exists(path))
        check("PDF > 0 bytes", os.path.getsize(path) > 0)
        print(f"    Path: {path}")

    # ── 8. Export Sales Register Excel ──
    print("\n[8] Export Sales Register Excel")
    ok, path = export_sales_register_excel()
    check("Excel exported", ok, path if not ok else "")
    if ok:
        check("Excel file exists", os.path.exists(path))
        check("Excel > 0 bytes", os.path.getsize(path) > 0)
        from openpyxl import load_workbook
        wb = load_workbook(path)
        check("Has Sales Register sheet", 'Sales Register' in wb.sheetnames, f"Sheets: {wb.sheetnames}")
        ws = wb['Sales Register']
        check("Has data rows", ws.max_row > 1, f"Rows: {ws.max_row}")
        wb.close()
        print(f"    Path: {path}")

    # ── Results ──
    print("\n" + "=" * 60)
    total = PASS + FAIL
    print(f"RESULTS: {PASS}/{total} passed, {FAIL} failed")
    if FAIL == 0:
        print("*** ALL TESTS PASSED -- Phase 6 is COMPLETE! ***")
    else:
        print("!!! Some tests failed -- review above.")
    print("=" * 60)
    return FAIL == 0

if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)
