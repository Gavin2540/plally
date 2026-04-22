"""
Phase 5 Integration Test -- PlywoodPro GST Reports
Runs after Phase 2 (sales invoice IGST to Chennai, purchase invoice IGST from Greenply).
"""
import os, sys
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

from modules.gst import (
    get_gstr1_b2b, get_gstr1_hsn_summary, get_gstr3b_summary, get_itc_register,
)
from utils.excel_export import export_gstr1_excel, export_itc_register_excel

PASS = FAIL = 0

def check(label, cond, detail=""):
    global PASS, FAIL
    if cond:
        PASS += 1; print(f"  [PASS] {label}")
    else:
        FAIL += 1; print(f"  [FAIL] {label} -- {detail}")

def main():
    global PASS, FAIL
    print("=" * 60)
    print("PlywoodPro Phase 5 -- GST Reports Integration Test")
    print("=" * 60)

    # Phase 2 test data uses date from test — need to find the month/year
    from db.connection import get_connection
    conn = get_connection()
    si = conn.execute(
        "SELECT date FROM vouchers WHERE voucher_type='Sales Invoice' AND status='Confirmed' LIMIT 1"
    ).fetchone()
    pi = conn.execute(
        "SELECT date FROM vouchers WHERE voucher_type='Purchase Invoice' AND status='Confirmed' LIMIT 1"
    ).fetchone()
    conn.close()

    if not si or not pi:
        print("  [FAIL] No Phase 2 invoices found. Run test_phase2.py first.")
        return False

    si_date = si['date']  # e.g. '2026-04-22'
    parts = si_date.split('-')
    month = int(parts[1]); year = int(parts[0])
    print(f"    Using month={month}, year={year} from invoice date {si_date}")

    # ── 1. GSTR-1 B2B ──
    print(f"\n[1] GSTR-1 B2B -- Month {month}/{year}")
    b2b = get_gstr1_b2b(month, year)
    check("B2B has entries", len(b2b) > 0, f"Got {len(b2b)}")
    if b2b:
        r = b2b[0]
        check("Party is Chennai Woodworks", 'Chennai' in r.get('party_name', ''), r.get('party_name'))
        check("IGST amount > 0", (r.get('igst_amount') or 0) > 0, f"IGST={r.get('igst_amount')}")
        check("Taxable value > 0", (r.get('taxable_amount') or 0) > 0, f"Taxable={r.get('taxable_amount')}")
        print(f"    GSTIN: {r.get('gstin')}, Taxable: {r.get('taxable_amount')}, IGST: {r.get('igst_amount')}")

    # ── 2. HSN Summary ──
    print(f"\n[2] GSTR-1 HSN Summary")
    hsn = get_gstr1_hsn_summary(month, year)
    check("HSN summary has entries", len(hsn) > 0, f"Got {len(hsn)}")
    hsn_4412 = next((h for h in hsn if h.get('hsn_code') == '4412'), None)
    check("HSN 4412 present", hsn_4412 is not None)
    if hsn_4412:
        check("HSN 4412 qty > 0", (hsn_4412.get('total_qty') or 0) > 0)
        check("HSN 4412 taxable > 0", (hsn_4412.get('total_taxable') or 0) > 0)
        print(f"    HSN 4412: qty={hsn_4412.get('total_qty')}, taxable={hsn_4412.get('total_taxable')}, igst={hsn_4412.get('igst')}")

    # ── 3. GSTR-3B ──
    print(f"\n[3] GSTR-3B Summary")
    g3b = get_gstr3b_summary(month, year)
    o = g3b.get('outward', {}); i = g3b.get('itc', {}); n = g3b.get('net_payable', {})
    check("Outward IGST > 0", (o.get('igst') or 0) > 0, f"Got {o.get('igst')}")
    check("ITC IGST > 0", (i.get('igst') or 0) > 0, f"Got {i.get('igst')}")
    net_igst = n.get('igst', 0)
    expected_net = (o.get('igst', 0) or 0) - (i.get('igst', 0) or 0)
    check("Net IGST = Payable - Input", abs(net_igst - expected_net) < 0.02,
          f"Net={net_igst}, Expected={expected_net}")
    print(f"    Outward IGST: {o.get('igst')}, ITC IGST: {i.get('igst')}, Net: {net_igst}")

    # ── 4. ITC Register ──
    print(f"\n[4] ITC Register")
    itc = get_itc_register()
    check("ITC register has entries", len(itc['entries']) > 0, f"Got {len(itc['entries'])}")
    check("Total eligible IGST = 4680", abs(itc['total_eligible_igst'] - 4680.0) < 0.02,
          f"Got {itc['total_eligible_igst']}")
    if itc['entries']:
        e = itc['entries'][0]
        check("First entry is Eligible", e.get('status') == 'Eligible', e.get('status'))
    print(f"    Eligible ITC: CGST={itc['total_eligible_cgst']}, SGST={itc['total_eligible_sgst']}, IGST={itc['total_eligible_igst']}")

    # ── 5. Excel Export ──
    print(f"\n[5] GSTR-1 Excel Export")
    ok, path = export_gstr1_excel(month, year)
    check("GSTR-1 Excel exported", ok, path if not ok else "")
    if ok:
        check("Excel file exists", os.path.exists(path))
        size = os.path.getsize(path) if os.path.exists(path) else 0
        check("Excel file > 0 bytes", size > 0, f"{size} bytes")

        # Check it has 4 sheets
        from openpyxl import load_workbook
        wb = load_workbook(path)
        sheets = wb.sheetnames
        check("Has 4 sheets", len(sheets) == 4, f"Got {sheets}")
        wb.close()
        print(f"    Sheets: {sheets}")
        print(f"    Path: {path}")

    # ── 6. GSTR-3B PDF Export ──
    print(f"\n[6] GSTR-3B PDF Export")
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from pathlib import Path

    export_dir = Path(PROJECT_ROOT) / "exports"
    export_dir.mkdir(exist_ok=True)
    pdf_path = str(export_dir / f"GSTR3B_{year}_{month:02d}.pdf")
    styles = getSampleStyleSheet()
    story = [Paragraph(f"GSTR-3B Summary -- {month:02d}/{year}", styles['Title']), Spacer(1, 5*mm)]
    tbl = [['Description','CGST','SGST','IGST','Total'],
           ['Outward', o.get('cgst',0), o.get('sgst',0), o.get('igst',0), sum(o.get(k,0) for k in ('cgst','sgst','igst'))],
           ['ITC', i.get('cgst',0), i.get('sgst',0), i.get('igst',0), sum(i.get(k,0) for k in ('cgst','sgst','igst'))],
           ['NET', n.get('cgst',0), n.get('sgst',0), n.get('igst',0), n.get('total',0)]]
    t = Table(tbl, colWidths=[140,80,80,80,90])
    t.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#2E7D32')),
        ('TEXTCOLOR',(0,0),(-1,0),colors.white),('GRID',(0,0),(-1,-1),0.3,colors.grey),
        ('ALIGN',(1,0),(-1,-1),'RIGHT'),('FONTNAME',(0,-1),(-1,-1),'Helvetica-Bold')]))
    story.append(t)
    doc = SimpleDocTemplate(pdf_path, pagesize=A4, leftMargin=15*mm, rightMargin=15*mm)
    doc.build(story)
    check("GSTR-3B PDF created", os.path.exists(pdf_path))
    check("PDF size > 0", os.path.getsize(pdf_path) > 0, f"{os.path.getsize(pdf_path)} bytes")
    print(f"    Path: {pdf_path}")

    # ── Results ──
    print("\n" + "=" * 60)
    total = PASS + FAIL
    print(f"RESULTS: {PASS}/{total} passed, {FAIL} failed")
    if FAIL == 0:
        print("*** ALL TESTS PASSED -- Phase 5 is COMPLETE! ***")
    else:
        print("!!! Some tests failed -- review above.")
    print("=" * 60)
    return FAIL == 0

if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)
